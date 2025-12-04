"""
Unit tests for extract_transform_platform.reports.excel_generator module.

Tests:
- ExcelReportGenerator initialization
- DataFrame conversion from various formats
- Report generation with different configurations
- Formatting application
- Excel features (freeze panes, auto-filter)
- Summary sheet generation
- Error handling
"""

from pathlib import Path

import pandas as pd
import pytest
from openpyxl import load_workbook

from extract_transform_platform.reports import (
    ExcelReportConfig,
    ExcelReportGenerator,
)


class TestExcelReportGeneratorInit:
    """Test ExcelReportGenerator initialization."""

    def test_init_sets_supported_features(self):
        """Test that initialization sets supported features."""
        generator = ExcelReportGenerator()

        features = generator.get_supported_features()
        assert "tables" in features
        assert "formatting" in features
        assert "freeze_panes" in features
        assert "auto_filter" in features


class TestDataFrameConversion:
    """Test _to_dataframe method."""

    @pytest.fixture
    def generator(self):
        """Create generator instance."""
        return ExcelReportGenerator()

    def test_dataframe_passthrough(self, generator):
        """Test that DataFrame is returned as-is."""
        df = pd.DataFrame({"a": [1, 2], "b": [3, 4]})
        result = generator._to_dataframe(df)

        assert result is df  # Same object
        assert result.equals(df)

    def test_dict_single_row(self, generator):
        """Test conversion from dict (single row)."""
        data = {"name": "Alice", "age": 30, "city": "NYC"}
        df = generator._to_dataframe(data)

        assert len(df) == 1
        assert list(df.columns) == ["name", "age", "city"]
        assert df.iloc[0]["name"] == "Alice"
        assert df.iloc[0]["age"] == 30

    def test_dict_columnar_format(self, generator):
        """Test conversion from dict of lists (columnar format)."""
        data = {"name": ["Alice", "Bob"], "age": [30, 25]}
        df = generator._to_dataframe(data)

        assert len(df) == 2
        assert list(df.columns) == ["name", "age"]
        assert df.iloc[0]["name"] == "Alice"
        assert df.iloc[1]["age"] == 25

    def test_list_of_dicts(self, generator):
        """Test conversion from list of dicts."""
        data = [
            {"name": "Alice", "age": 30},
            {"name": "Bob", "age": 25},
        ]
        df = generator._to_dataframe(data)

        assert len(df) == 2
        assert list(df.columns) == ["name", "age"]
        assert df.iloc[0]["name"] == "Alice"

    def test_list_of_lists(self, generator):
        """Test conversion from list of lists."""
        data = [["Alice", 30], ["Bob", 25]]
        df = generator._to_dataframe(data)

        assert len(df) == 2
        assert len(df.columns) == 2  # Auto-generated column names

    def test_list_of_values(self, generator):
        """Test conversion from list of values (single column)."""
        data = [1, 2, 3, 4, 5]
        df = generator._to_dataframe(data)

        assert len(df) == 5
        assert "value" in df.columns

    def test_empty_list(self, generator):
        """Test conversion from empty list."""
        data = []
        df = generator._to_dataframe(data)

        assert len(df) == 0
        assert isinstance(df, pd.DataFrame)

    def test_unsupported_type_raises(self, generator):
        """Test that unsupported type raises TypeError."""
        with pytest.raises(TypeError, match="Unsupported data type"):
            generator._to_dataframe(42)

        with pytest.raises(TypeError, match="Unsupported data type"):
            generator._to_dataframe("string")


class TestReportGeneration:
    """Test generate method."""

    @pytest.fixture
    def generator(self):
        """Create generator instance."""
        return ExcelReportGenerator()

    @pytest.fixture
    def sample_data(self):
        """Create sample DataFrame."""
        return pd.DataFrame(
            {
                "name": ["Alice", "Bob", "Charlie"],
                "age": [30, 25, 35],
                "city": ["NYC", "LA", "SF"],
            }
        )

    @pytest.fixture
    def output_path(self, tmp_path):
        """Create output path."""
        return tmp_path / "test_report.xlsx"

    def test_generate_basic_report(self, generator, sample_data, output_path):
        """Test basic report generation."""
        config = ExcelReportConfig(title="Test Report")

        result_path = generator.generate(sample_data, output_path, config)

        # File should exist
        assert result_path.exists()
        assert result_path == output_path

        # Verify file is valid Excel
        wb = load_workbook(result_path)
        assert "Report" in wb.sheetnames

        # Read back data
        df_read = pd.read_excel(result_path, sheet_name="Report", skiprows=4)
        assert len(df_read) == 3
        assert list(df_read.columns) == ["name", "age", "city"]

    def test_generate_with_dict_data(self, generator, output_path):
        """Test generation with dict input."""
        data = {"name": "Alice", "age": 30}
        config = ExcelReportConfig(title="Dict Report")

        result_path = generator.generate(data, output_path, config)
        assert result_path.exists()

        df_read = pd.read_excel(result_path, sheet_name="Report", skiprows=4)
        assert len(df_read) == 1

    def test_generate_with_list_data(self, generator, output_path):
        """Test generation with list input."""
        data = [
            {"name": "Alice", "age": 30},
            {"name": "Bob", "age": 25},
        ]
        config = ExcelReportConfig(title="List Report")

        result_path = generator.generate(data, output_path, config)
        assert result_path.exists()

        df_read = pd.read_excel(result_path, sheet_name="Report", skiprows=4)
        assert len(df_read) == 2

    def test_generate_without_timestamp(self, generator, sample_data, output_path):
        """Test generation without timestamp metadata."""
        config = ExcelReportConfig(title="No Timestamp Report", include_timestamp=False)

        result_path = generator.generate(sample_data, output_path, config)
        assert result_path.exists()

        # Data should start at row 1 (no metadata)
        df_read = pd.read_excel(result_path, sheet_name="Report")
        assert len(df_read) == 3

    def test_generate_with_custom_sheet_name(self, generator, sample_data, output_path):
        """Test generation with custom sheet name."""
        config = ExcelReportConfig(title="Custom Sheet", sheet_name="Sales Data")

        result_path = generator.generate(sample_data, output_path, config)

        wb = load_workbook(result_path)
        assert "Sales Data" in wb.sheetnames

    def test_generate_with_summary_sheet(self, generator, sample_data, output_path):
        """Test generation with summary sheet."""
        config = ExcelReportConfig(title="Report with Summary", include_summary=True)

        result_path = generator.generate(sample_data, output_path, config)

        wb = load_workbook(result_path)
        assert "Summary" in wb.sheetnames
        assert "Report" in wb.sheetnames

    def test_generate_with_custom_column_widths(
        self, generator, sample_data, output_path
    ):
        """Test generation with custom column widths."""
        config = ExcelReportConfig(
            title="Custom Widths",
            column_widths={"name": 25, "age": 10, "city": 20},
        )

        result_path = generator.generate(sample_data, output_path, config)

        wb = load_workbook(result_path)
        ws = wb["Report"]

        # Check column widths (approximately, openpyxl may adjust slightly)
        name_width = ws.column_dimensions["A"].width
        age_width = ws.column_dimensions["B"].width
        city_width = ws.column_dimensions["C"].width

        assert name_width == pytest.approx(25, abs=2)
        assert age_width == pytest.approx(10, abs=2)
        assert city_width == pytest.approx(20, abs=2)

    def test_generate_returns_absolute_path(self, generator, sample_data, output_path):
        """Test that generate returns absolute path."""
        config = ExcelReportConfig(title="Test")

        result_path = generator.generate(sample_data, output_path, config)

        assert result_path.is_absolute()


class TestFormatting:
    """Test formatting application."""

    @pytest.fixture
    def generator(self):
        """Create generator instance."""
        return ExcelReportGenerator()

    @pytest.fixture
    def sample_data(self):
        """Create sample DataFrame."""
        return pd.DataFrame(
            {
                "name": ["Alice", "Bob"],
                "age": [30, 25],
            }
        )

    @pytest.fixture
    def output_path(self, tmp_path):
        """Create output path."""
        return tmp_path / "test_report.xlsx"

    def test_header_row_formatting(self, generator, sample_data, output_path):
        """Test that header row has correct formatting."""
        config = ExcelReportConfig(title="Test")
        generator.generate(sample_data, output_path, config)

        wb = load_workbook(output_path)
        ws = wb["Report"]

        # Header row is at row 5 (after metadata)
        header_cell = ws.cell(row=5, column=1)

        # Check bold font
        assert header_cell.font.bold is True

        # Check white text color (ARGB format: alpha + RGB)
        assert header_cell.font.color.rgb in [
            "FFFFFFFF",
            "00FFFFFF",
        ]  # Allow both ARGB formats

        # Check blue background (ARGB format)
        assert header_cell.fill.start_color.rgb in [
            "FF366092",
            "00366092",
        ]  # Allow both ARGB formats


class TestExcelFeatures:
    """Test Excel-specific features."""

    @pytest.fixture
    def generator(self):
        """Create generator instance."""
        return ExcelReportGenerator()

    @pytest.fixture
    def sample_data(self):
        """Create sample DataFrame with more rows."""
        return pd.DataFrame(
            {
                "name": [f"Person {i}" for i in range(20)],
                "age": list(range(20, 40)),
            }
        )

    @pytest.fixture
    def output_path(self, tmp_path):
        """Create output path."""
        return tmp_path / "test_report.xlsx"

    def test_freeze_panes_enabled(self, generator, sample_data, output_path):
        """Test that freeze panes is applied when enabled."""
        config = ExcelReportConfig(title="Freeze Test", freeze_header=True)
        generator.generate(sample_data, output_path, config)

        wb = load_workbook(output_path)
        ws = wb["Report"]

        # Freeze panes should be set
        assert ws.freeze_panes is not None

    def test_freeze_panes_disabled(self, generator, sample_data, output_path):
        """Test that freeze panes is not applied when disabled."""
        config = ExcelReportConfig(title="No Freeze", freeze_header=False)
        generator.generate(sample_data, output_path, config)

        wb = load_workbook(output_path)
        ws = wb["Report"]

        # Freeze panes should not be set
        assert ws.freeze_panes is None

    def test_auto_filter_enabled(self, generator, sample_data, output_path):
        """Test that auto-filter is applied when enabled."""
        config = ExcelReportConfig(title="Filter Test", auto_filter=True)
        generator.generate(sample_data, output_path, config)

        wb = load_workbook(output_path)
        ws = wb["Report"]

        # Auto-filter should be set
        assert ws.auto_filter.ref is not None

    def test_auto_filter_disabled(self, generator, sample_data, output_path):
        """Test that auto-filter is not applied when disabled."""
        config = ExcelReportConfig(title="No Filter", auto_filter=False)
        generator.generate(sample_data, output_path, config)

        wb = load_workbook(output_path)
        ws = wb["Report"]

        # Auto-filter should not be set (empty ref)
        assert ws.auto_filter.ref is None


class TestErrorHandling:
    """Test error handling."""

    @pytest.fixture
    def generator(self):
        """Create generator instance."""
        return ExcelReportGenerator()

    def test_invalid_extension_raises(self, generator, tmp_path):
        """Test that wrong file extension raises ValueError."""
        output_path = tmp_path / "report.txt"
        data = pd.DataFrame({"a": [1]})
        config = ExcelReportConfig(title="Test")

        with pytest.raises(ValueError, match="must have .xlsx extension"):
            generator.generate(data, output_path, config)

    def test_none_data_raises(self, generator, tmp_path):
        """Test that None data raises ValueError."""
        output_path = tmp_path / "report.xlsx"
        config = ExcelReportConfig(title="Test")

        with pytest.raises(ValueError, match="cannot be None"):
            generator.generate(None, output_path, config)

    def test_empty_data_raises(self, generator, tmp_path):
        """Test that empty data raises ValueError."""
        output_path = tmp_path / "report.xlsx"
        config = ExcelReportConfig(title="Test")

        with pytest.raises(ValueError, match="cannot be empty"):
            generator.generate([], output_path, config)

    def test_unsupported_data_type_raises(self, generator, tmp_path):
        """Test that unsupported data type raises TypeError."""
        output_path = tmp_path / "report.xlsx"
        config = ExcelReportConfig(title="Test")

        with pytest.raises(TypeError, match="Unsupported data type"):
            generator.generate(42, output_path, config)
